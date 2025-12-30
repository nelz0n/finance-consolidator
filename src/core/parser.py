"""File parser for different institution formats."""

import csv
import re
from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl
from src.utils.logger import get_logger

logger = get_logger()


class FileParser:
    """Parse CSV/XLSX files based on institution configuration."""

    def __init__(self, institution_config: Dict[str, Any]):
        """
        Initialize parser with institution configuration.

        Args:
            institution_config: Institution configuration from YAML
        """
        self.config = institution_config
        self.institution_name = institution_config.get('institution', {}).get('name', 'Unknown')

        # Support both 'format' (new) and 'csv_format' (old) for backward compatibility
        self.format_config = institution_config.get('format', institution_config.get('csv_format', {}))

        # Determine file type
        file_type = self.format_config.get('type') or self.format_config.get('file_type', 'csv')
        # Handle special case: "xlsx_with_csv" → xlsx
        if 'xlsx' in file_type.lower():
            self.file_type = 'xlsx'
        else:
            self.file_type = file_type

        logger.debug(f"Initialized parser for {self.institution_name} (format: {self.file_type})")

    def parse_file(self, file_path: str, original_filename: str = None) -> List[Dict[str, Any]]:
        """
        Parse file based on institution configuration.

        Args:
            file_path: Path to file to parse
            original_filename: Optional original filename (for account extraction when file was renamed)

        Returns:
            List of raw transaction dictionaries
        """
        logger.info(f"Parsing file: {file_path}")

        file_path_obj = Path(file_path)

        # Use original filename for account extraction if provided, otherwise use actual file path
        filename_for_extraction = original_filename if original_filename else file_path

        try:
            # Use format.type from config (fallback to file extension)
            file_type = self.file_type
            if not file_type:
                file_extension = file_path_obj.suffix.lower()
                file_type = 'xlsx' if file_extension in ['.xlsx', '.xls'] else 'csv'

            # Route to appropriate parser
            if file_type == 'csv':
                transactions = self._parse_csv(file_path, filename_for_extraction)
            elif file_type == 'xlsx':
                transactions = self._parse_xlsx(file_path, filename_for_extraction)
            else:
                logger.error(f"Unsupported file type: {file_type}")
                return []

            logger.info(f"Successfully parsed {len(transactions)} transactions from {file_path_obj.name}")
            return transactions

        except Exception as e:
            logger.error(f"Error parsing file {file_path}: {str(e)}")
            return []

    def _parse_csv(self, file_path: str, filename_for_extraction: str = None) -> List[Dict[str, Any]]:
        """
        Parse CSV file using configuration.

        Args:
            file_path: Path to CSV file
            filename_for_extraction: Optional filename for account extraction (if file was renamed)

        Returns:
            List of transaction dictionaries
        """
        logger.debug(f"Parsing CSV: {file_path}")

        # Use filename for extraction (defaults to file_path if not provided)
        if not filename_for_extraction:
            filename_for_extraction = file_path

        transactions = []

        # Get format settings
        encoding = self.format_config.get('encoding', 'utf-8')
        delimiter = self.format_config.get('delimiter', ',')
        skip_rows = self.format_config.get('skip_rows', 0)
        has_header = self.format_config.get('has_header', True)

        # Get column mapping and filtering
        column_mapping = self.config.get('columns', self.config.get('column_mapping', {}))
        filtering = self.config.get('filtering', {})
        skip_patterns = filtering.get('skip_if_contains', [])

        try:
            with open(file_path, 'r', encoding=encoding) as f:
                # Skip rows if needed
                for _ in range(skip_rows):
                    next(f, None)

                # Read CSV with or without header
                if has_header:
                    reader = csv.DictReader(f, delimiter=delimiter)
                else:
                    # For headerless CSV, use numeric indices
                    reader = csv.reader(f, delimiter=delimiter)
                    reader = [dict(enumerate(row)) for row in reader]

                for row_num, row in enumerate(reader, start=skip_rows + 1):
                    # Skip rows that match filtering patterns
                    skip_row = False
                    for pattern in skip_patterns:
                        row_values = row.values() if isinstance(row, dict) else row
                        if any(pattern in str(value) for value in row_values):
                            logger.debug(f"Skipping row {row_num} (matches pattern: {pattern})")
                            skip_row = True
                            break

                    if skip_row:
                        continue

                    # Apply transformations if defined
                    if 'transformations' in self.config:
                        row = self._apply_transformations(row)

                    # Map columns to standard fields
                    transaction = self._map_columns(row, column_mapping, filename_for_extraction)

                    if transaction:
                        transactions.append(transaction)

        except Exception as e:
            logger.error(f"Error parsing CSV: {str(e)}")
            raise

        return transactions

    def _parse_xlsx(self, file_path: str, filename_for_extraction: str = None) -> List[Dict[str, Any]]:
        """
        Parse XLSX file using configuration.

        Args:
            file_path: Path to XLSX file
            filename_for_extraction: Optional filename for account extraction (if file was renamed)

        Returns:
            List of transaction dictionaries
        """
        logger.debug(f"Parsing XLSX: {file_path}")

        # Use filename for extraction (defaults to file_path if not provided)
        if not filename_for_extraction:
            filename_for_extraction = file_path

        transactions = []

        # Get format settings
        sheet_name = self.format_config.get('sheet_name', None)  # None = first sheet
        skip_rows = self.format_config.get('skip_rows', 0)

        column_mapping = self.config.get('columns', self.config.get('column_mapping', {}))

        try:
            # Load workbook
            wb = openpyxl.load_workbook(file_path, data_only=True)

            # Get sheet
            if sheet_name:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            # Check if we need special Partners Bank handling (concatenate columns A-D)
            # This is indicated by merged_columns: true in the config
            use_partners_logic = (
                self.format_config.get('merged_columns', False) or
                'Partners' in self.institution_name
            )

            if use_partners_logic:
                # Use Partners-specific logic for backward compatibility
                transactions = self._parse_partners_xlsx_logic(ws, column_mapping, file_path)
            else:
                # Generic XLSX parsing
                # Read header from first row (after skip_rows)
                header_row = skip_rows + 1
                header = []

                for col_idx, cell in enumerate(ws[header_row]):
                    if cell.value is not None:
                        header.append(str(cell.value).strip())
                    else:
                        header.append(f"col_{col_idx}")

                logger.debug(f"XLSX header: {len(header)} columns")

                # Parse data rows
                for row_num in range(header_row + 1, ws.max_row + 1):
                    row_dict = {}

                    # Build dict with both column names and indices
                    for col_idx, cell in enumerate(ws[row_num]):
                        value = cell.value
                        if value is not None:
                            value_str = str(value).strip()
                        else:
                            value_str = ''

                        # Add by column name
                        if col_idx < len(header):
                            row_dict[header[col_idx]] = value_str

                        # Also add by index
                        row_dict[col_idx] = value_str

                    # Skip empty rows
                    if not any(row_dict.values()):
                        continue

                    # Apply transformations if defined
                    if 'transformations' in self.config:
                        row_dict = self._apply_transformations(row_dict)

                    # Map columns to standard fields
                    transaction = self._map_columns(row_dict, column_mapping, filename_for_extraction)

                    if transaction:
                        transactions.append(transaction)

            wb.close()

        except Exception as e:
            logger.error(f"Error parsing XLSX: {str(e)}")
            raise

        return transactions

    def _parse_partners_xlsx_logic(self, ws, column_mapping: Dict, file_path: str) -> List[Dict[str, Any]]:
        """
        Special Partners Bank XLSX parsing logic.

        Concatenates columns A-D and splits by semicolon.
        Kept for backward compatibility.
        """
        transactions = []

        # Extract account number from filename
        account_number = self._extract_account_from_filename(Path(file_path).name)
        if account_number:
            logger.debug(f"Extracted account number: {account_number}")

        # Build header from row 1
        header_parts = []
        for col in ['A', 'B', 'C', 'D']:
            cell_value = ws[f'{col}1'].value
            if cell_value is not None:
                header_parts.append(str(cell_value))

        header_string = ''.join(header_parts)
        header = header_string.split(';')
        logger.debug(f"Partners header fields: {len(header)} columns")

        # Parse data rows (starting from row 2)
        for row_num in range(2, ws.max_row + 1):
            # Concatenate columns A, B, C, D
            parts = []
            for col in ['A', 'B', 'C', 'D']:
                cell_value = ws[f'{col}{row_num}'].value
                if cell_value is not None:
                    parts.append(str(cell_value))

            if not parts:
                continue

            # Join all parts and split by semicolon
            full_row = ''.join(parts)
            fields = full_row.split(';')

            # Create dict with both index and name access
            row_dict = {}

            # Add by header name
            for i, field_name in enumerate(header):
                if i < len(fields):
                    row_dict[field_name.strip()] = fields[i].strip()

            # Also add by index for compatibility
            for i, field_value in enumerate(fields):
                row_dict[i] = field_value.strip()

            # Add account number from filename
            if account_number:
                row_dict['account'] = account_number

            # Map columns to standard fields
            transaction = self._map_columns(row_dict, column_mapping, file_path)

            if transaction:
                transactions.append(transaction)

        return transactions

    def _apply_transformations(self, row: Dict[str, Any]) -> Dict[str, Any]:
        """
        Apply transformations defined in config.

        Supports:
        - concatenate: Join multiple columns
        - strip: Remove characters
        - replace: Replace text
        - split: Split by delimiter

        Args:
            row: Row dictionary

        Returns:
            Transformed row dictionary
        """
        transformations = self.config.get('transformations', {})

        for column, transform_expr in transformations.items():
            try:
                # Handle concatenation: "A + B + C" or "8 + ' [Msg: ' + 4"
                if '+' in transform_expr:
                    parts = [p.strip().strip("'\"") for p in transform_expr.split('+')]
                    values = []
                    for part in parts:
                        # Try as integer index first (for Partners Bank numeric indices)
                        try:
                            idx = int(part)
                            if idx in row:
                                values.append(str(row[idx]))
                                continue
                        except ValueError:
                            pass

                        # Then try as string key (for column names)
                        if part in row:
                            values.append(str(row[part]))
                        else:
                            values.append(part)  # Literal string
                    row[column] = ''.join(values)

                # Handle strip: "strip('xyz')"
                elif 'strip(' in transform_expr:
                    match = re.search(r"strip\(['\"](.+?)['\"]\)", transform_expr)
                    if match:
                        chars = match.group(1)
                        if column in row:
                            row[column] = str(row[column]).strip(chars)

                # Handle replace: "replace('old', 'new')"
                elif 'replace(' in transform_expr:
                    match = re.search(r"replace\(['\"](.+?)['\"]\s*,\s*['\"](.+?)['\"]\)", transform_expr)
                    if match:
                        old_text = match.group(1)
                        new_text = match.group(2)
                        if column in row:
                            row[column] = str(row[column]).replace(old_text, new_text)

                # Handle split: "split(';')[0]"
                elif 'split(' in transform_expr:
                    match = re.search(r"split\(['\"](.+?)['\"]\)\[(\d+)\]", transform_expr)
                    if match:
                        delimiter = match.group(1)
                        index = int(match.group(2))
                        if column in row:
                            parts = str(row[column]).split(delimiter)
                            if index < len(parts):
                                row[column] = parts[index].strip()

            except Exception as e:
                logger.warning(f"Error applying transformation to column '{column}': {str(e)}")

        return row

    def _map_columns(self, row: Dict[str, Any], column_mapping: Dict[str, Any], file_path: str = None) -> Optional[Dict[str, Any]]:
        """
        Map CSV columns to standard transaction fields.

        Args:
            row: Raw row dictionary from CSV
            column_mapping: Column mapping from config
            file_path: Optional file path for extract_from_filename feature

        Returns:
            Mapped transaction dictionary or None if invalid
        """
        transaction = {}

        # Map each field
        for standard_field, csv_column in column_mapping.items():
            if standard_field == 'defaults':
                # Handle defaults separately
                continue

            # Get value from row
            if isinstance(csv_column, str):
                # String column name
                value = row.get(csv_column, '')
            elif isinstance(csv_column, int):
                # Integer index (for Partners Bank)
                value = row.get(csv_column, '')
            else:
                # Direct value
                value = csv_column

            transaction[standard_field] = value

        # Add defaults
        defaults = column_mapping.get('defaults', {})
        for field, default_value in defaults.items():
            if field not in transaction or not transaction[field]:
                # Handle special "extract_from_filename" directive
                if default_value == "extract_from_filename" and file_path:
                    extracted_value = self._extract_account_from_filename(Path(file_path).name)
                    # Append bank code if configured
                    if extracted_value and 'account_bank_code' in defaults:
                        bank_code = defaults['account_bank_code']
                        extracted_value = f"{extracted_value}/{bank_code}"
                    transaction[field] = extracted_value if extracted_value else ''
                else:
                    transaction[field] = default_value

        # Skip empty transactions
        if not transaction.get('date') and not transaction.get('amount'):
            return None

        return transaction

    def _extract_account_from_filename(self, filename: str) -> Optional[str]:
        r"""
        Extract account number from Partners Bank filename.

        Format: vypis_<account>_<dates>.xlsx
        Example: vypis_1330299329_20251001_20251031.xlsx

        Args:
            filename: Filename to parse

        Returns:
            Account number or None
        """
        pattern = r'vypis_(\d+)_'
        match = re.search(pattern, filename)

        if match:
            return match.group(1)

        logger.warning(f"Could not extract account number from filename: {filename}")
        return None
